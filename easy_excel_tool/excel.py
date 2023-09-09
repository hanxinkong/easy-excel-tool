import os
from typing import Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

__author__ = 'hanxinkong'
__author_email__ = 'xinkonghan@gmail.com'


class MyException(Exception):
    def __init__(self, *args):
        self.args = args


class CreateError(MyException):
    def __init__(self, code=100, message='Create Exception', args=('Create Exception',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class OpenError(MyException):
    def __init__(self, code=101, message='Open Exception', args=('Open Exception',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class ParamError(MyException):
    def __init__(self, code=102, message='Parameter exception', args=('Parameter exception',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class ParamTypeError(MyException):
    def __init__(self, code=103, message='Wrong parameter type', args=('Wrong parameter type',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class FileExistenceError(MyException):
    def __init__(self, code=104, message='The file already exists', args=('The file already exists',)):
        self.args = args
        self.message = message
        self.code = code

    def __str__(self):
        return self.message


class Excel(object):
    def __init__(self, file: str):
        """
        This is a very convenient tool for operating excel.
        After the secondary encapsulation of excel operation by Pandas,
        The purpose of this tool is to make it easier without focusing on the process.
        """
        self.__wb = None
        self.file = file

    def create_excel(self, columns: list = None, sheet_name: str = 'Sheet1', inplace: bool = False) -> None:
        """
        Create an excel document
        :param sheet_name:
        :param inplace: Overwrite if the file exists
        :param columns: If it is blank, a new blank Excel document will be created
        :return:
        """
        if os.path.isfile(self.file):
            """
            Overwrite or rebuild files
            """
            if isinstance(inplace, bool) and inplace is False:
                """
                Keep the original file
                """
                raise FileExistenceError(
                    message=f'The {self.file} file already exists, please check whether to replace it, change the '
                            f'parameter inplace=True, or delete this file manually')

        result = pd.DataFrame(columns=columns)
        try:
            result.to_excel(self.file, sheet_name=sheet_name, engine='openpyxl')
        except Exception:
            raise CreateError

    def add_sheet(self, columns: list = None, sheet_name=None, inplace: bool = False):
        """

        :param columns:
        :param sheet_name:
        :param inplace:
        :return:
        """
        if sheet_name is None:
            sheet_name = []
        # result = pd.DataFrame(columns=columns)
        if os.path.isfile(self.file):
            for title in sheet_name:
                try:
                    self.__wb.create_sheet(
                        title=title,
                        # index=0
                    )
                    # with pd.ExcelWriter(self.file, mode='a', engine='openpyxl') as wf:
                    #     result.to_excel(wf, index=False, header=False, sheet_name=i)
                except Exception:
                    raise CreateError

    def get_sheet_name(self, excel_file: str) -> list:
        if os.path.isfile(excel_file):
            if self.__wb is not None:
                sheet_names = self.__wb.sheetnames
                return sheet_names
        return []

    def remove_sheet(self, sheet_name: str):
        if self.__wb is not None:
            ws = self.__wb[sheet_name]
            self.__wb.remove(ws)
            self.__wb.save(self.file)

    def write_excel(
            self,
            data: list,
            columns: list = None,
            sheet_name: str = 'Sheet1',
            mode: str = 'w',
            fill_column: Dict[str, Any] = None,
            header: bool = False,
            inplace: bool = False,
    ):
        """
        Write content to an excel file sheet page
        :param inplace:
        :param header:
        :param data:
        :param columns:
        :param sheet_name:
        :param mode: w 覆盖式写入内容 ；w+ 追加内容；a 向新的sheet页追加 ;str类型
        :param fill_column: The last element is used as the filling column，单独在最后追加一列；bool
        :return:
        """
        assert data, 'Data Parameter exception'

        if not self.file:
            raise ParamError(message=f'Check file path parameters')

        data = [{k: ILLEGAL_CHARACTERS_RE.sub(r'', str(v)) for k, v in i.items()} for i in data]

        excel_data = pd.DataFrame(data, columns=columns)

        if fill_column and isinstance(fill_column, dict):
            for k, v in fill_column.items():
                excel_data[k] = ILLEGAL_CHARACTERS_RE.sub(r'', str(v))

        '''
           If the file does not exist, create the document first
        '''
        if os.path.isfile(self.file) is False or inplace:
            self.create_excel(sheet_name=sheet_name, inplace=inplace)

        try:
            self.__wb = wb = load_workbook(self.file)

            '''
               Check whether the sheet name exists
               If it does not exist, the corresponding sheet will be created
            '''
            sheet_names = self.get_sheet_name(excel_file=self.file)
            exits_sheet_name = True if sheet_name in sheet_names else False

            if exits_sheet_name is True and mode == 'w':
                '''
                    First delete the old sheet, then create a new one
                '''
                if len(sheet_names) > 1:
                    self.remove_sheet(sheet_name=sheet_name)

            if exits_sheet_name is False and mode == 'w+':
                self.add_sheet(sheet_name=[sheet_name])

            if exits_sheet_name is True and mode == 'a':
                raise CreateError(
                    message='When you enable appending in mode a, you must ensure that the sheet does not exist')

            '''
                * When there is no content in the first row, insert the header, and no header will be added for subsequent appending
                Append data to the specified sheet
            '''
            sheet = wb.worksheets[wb.sheetnames.index(sheet_name)]
            rows = [row for row in sheet.rows]

            if header is False:
                exits_header = True
            else:
                exits_header = True if rows else False

            for row in dataframe_to_rows(
                    excel_data,
                    header=False if exits_header else True,
                    index=False,
            ):
                sheet.append(row)
            wb.save(self.file)
            wb.close()
        except Exception:
            raise OpenError
